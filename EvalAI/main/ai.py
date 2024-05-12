import os
from openpyxl import Workbook, load_workbook
from .ai12 import process_images, process_images1




def assign_grade(accuracy_percentage):
    if accuracy_percentage >= 0 and accuracy_percentage < 41:
        return 2
    elif accuracy_percentage >= 41 and accuracy_percentage < 61:
        return 3
    elif accuracy_percentage >= 61 and accuracy_percentage < 81:
        return 4
    elif accuracy_percentage >= 81 and accuracy_percentage <= 100:
        return 5
    else:
        return None


def get_image_files(directory):
    """Получить список файлов изображений в указанной директории."""
    image_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.bmp')
    image_files = [f for f in os.listdir(directory) if f.lower().endswith(image_extensions)]
    trot_directory = os.path.join(directory, 'trot')
    contents_of_trot = os.listdir(trot_directory)
    return image_files, contents_of_trot


def create_excel_with_images_from_template(input_dir, output_dir, excel_filename, zzz,  template_name="AREG.xlsx"):
    """Создать Excel файл на основе шаблона с именами изображений и общим количеством в указанной папке."""

    # Определение пути к шаблону и к новому файлу
    template_path = os.path.join(output_dir, template_name)
    output_path = os.path.join(output_dir, excel_filename)

# Загрузка книги шаблона и выбор активной страницы
    if os.path.exists(template_path):
        workbook = load_workbook(filename=template_path)
        sheet = workbook.active
        sheet['A2'] = zzz[0]
        sheet['B2'] = zzz[1]
        sheet['C2'] = zzz[2]


        # получение результатов в виде словаря(ответы)
        image_files, contents_of_trot = get_image_files(input_dir)  # списоки с названиями картин, и ответы тоже список
        print(image_files, contents_of_trot)
        file_path = os.path.join(input_dir, image_files[0])  # путь к 1 фотогорафии
        trot_dir = os.path.join(input_dir, 'trot')  # путь к троф
        file_path12 = os.path.join(trot_dir, contents_of_trot[0])  # путь к ответам
        # a, b, v = process_images(file_path, file_path12)
        x = process_images1(file_path12)  # словарь правильных ответов
        a = len(x)  # длина списка
        x.pop("name", None)
        print(x)

        current_row = 4  # Текущая строка для записи данных
        number_student = 0
        # получение в виде словаря(тест)
        for i in image_files:
            current_row += 1
            number_student += 1
            file_path = os.path.join(input_dir, i)  # путь к 1 фотогорафии
            y = process_images1(file_path)

            student_name = y['name']  # фио студента

            # Удаляем первый элемент (ФИО студента) из словарей
            y.pop('name', None)

            # Подсчет общего числа вопросов
            total_questions = a - 1
            sheet['D2'] = total_questions

            # Подсчет числа правильных ответов
            correct_count = sum(1 for key in y if x.get(key) is not None and y[key] == x[key])


            # Вычисление процента правильных ответов
            accuracy_percentage = (correct_count / total_questions) * 100 if total_questions > 0 else 0
            print(
                f"{accuracy_percentage} - % правильных ответов по отношению к другим ответам, студента {student_name}")

            # заполнение иксель таблицы
            sheet[f'A{current_row}'] = number_student
            sheet[f'B{current_row}'] = student_name
            sheet[f'C{current_row}'] = correct_count
            sheet[f'D{current_row}'] = accuracy_percentage
            sheet[f'E{current_row}'] = assign_grade(accuracy_percentage)


    else:
        #Если шаблон не найден, создаем новую книгу и страницу
       workbook = Workbook()
       sheet = workbook.active




    # Сохраняем измененную книгу как новый файл
    workbook.save(output_path)


